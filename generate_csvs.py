"""
CSV Constructions — SharePoint to CSV Pipeline
Downloads Excel files from SharePoint via Microsoft Graph API,
transforms data, and uploads resulting CSVs back to SharePoint.
"""

import os
import io
import csv
import requests
import openpyxl

# ─── SHAREPOINT / GRAPH CONFIG ───────────────────────────────────────────────
TENANT_ID     = os.environ["TENANT_ID"]
CLIENT_ID     = os.environ["CLIENT_ID"]
CLIENT_SECRET = os.environ["CLIENT_SECRET"]

# DataHub Shared Documents drive
DRIVE_ID = "b!_bzXaIx86kOufgJN3ih-BaDIDthKYuxJkJtLi1Bm5irGjCEnK-VHSpBRRm3_SDKU"


# ─── MICROSOFT GRAPH AUTH ─────────────────────────────────────────────────────

def get_access_token():
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    resp = requests.post(url, data={
        "grant_type":    "client_credentials",
        "client_id":     CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope":         "https://graph.microsoft.com/.default",
    })
    resp.raise_for_status()
    return resp.json()["access_token"]


# ─── SHAREPOINT FILE OPERATIONS ──────────────────────────────────────────────

def download_file_by_path(token, path):
    """Download a file from SharePoint by its path relative to the drive root."""
    url = (f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}"
           f"/root:/{path}:/content")
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    resp.raise_for_status()
    return io.BytesIO(resp.content)


def upload_file(token, path, content, content_type="text/csv"):
    """Upload a file to SharePoint by path relative to the drive root."""
    url = (f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}"
           f"/root:/{path}:/content")
    resp = requests.put(
        url,
        headers={"Authorization": f"Bearer {token}", "Content-Type": content_type},
        data=content,
    )
    resp.raise_for_status()
    print(f"Uploaded: {path}")


# ─── CSV BUILDERS ─────────────────────────────────────────────────────────────

def build_location_management(token):
    """
    Read the Locations table from List Source.xlsx.
    Create a CSV with columns [Location, Manager, LMS Location] containing
    every unique (Location, Manager) pair from RM I, RM II, and RD columns,
    excluding rows where the manager is blank.
    """
    print("Downloading List Source.xlsx...")
    excel_bytes = download_file_by_path(token, "Definitive Lists/List Source.xlsx")
    wb = openpyxl.load_workbook(excel_bytes, data_only=True)
    ws = wb["Locations"]

    # Build header index
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        location = row[col["Location"]]
        paychex = row[col["Paylocity Name"]]
        if not location:
            continue
        for mgr_col in ("RM I", "RM II", "RD"):
            manager = row[col[mgr_col]]
            if manager and str(manager).strip():
                rows.append((str(location).strip(),
                             str(manager).strip(),
                             str(paychex).strip() if paychex else ""))

    # Deduplicate while preserving order
    seen = set()
    unique_rows = []
    for r in rows:
        key = (r[0], r[1])
        if key not in seen:
            seen.add(key)
            unique_rows.append(r)

    # Write CSV to buffer
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Location", "Manager", "LMS Location"])
    writer.writerows(unique_rows)
    csv_bytes = buf.getvalue().encode("utf-8")

    print(f"Location Management: {len(unique_rows)} rows")
    upload_file(token, "Power BI Data Sources/Location Management.csv", csv_bytes)


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print("Authenticating with Microsoft Graph...")
    token = get_access_token()

    build_location_management(token)

    print("All done!")


if __name__ == "__main__":
    main()
